"""SAINATH AGRO INDUSTRIES - ERP main application."""
import io
import re
import csv
import json
from urllib.parse import urlparse
from datetime import date, timedelta, datetime

from flask import (
    Flask, render_template, request, redirect, url_for, jsonify,
    session, flash, Response, g, abort,
)
from werkzeug.security import check_password_hash, generate_password_hash

from models import get_conn, init_db
import auth
import validators
from auth import login_required, require, can_access, current_user
from auth import load_perms, template_for, ALL_MODULES, MODULE_KEYS
import logic
import backup as backup_mod
import os
import time
import hmac
import secrets
from functools import wraps as _wraps

app = Flask(__name__)
# Secret key: use env var in production; fall back to a generated one for dev.
app.secret_key = os.environ.get("SECRET_KEY") or "sainath-erp-change-this-in-production"

# ---- secure session cookie settings ----
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,      # JS cannot read the cookie (blocks XSS theft)
    SESSION_COOKIE_SAMESITE="Lax",     # blocks most cross-site request forgery
    SESSION_COOKIE_SECURE=bool(os.environ.get("HTTPS_ONLY", "1") == "1"),  # HTTPS only
    PERMANENT_SESSION_LIFETIME=60 * 60 * 12,   # auto-logout after 12h
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,       # reject uploads >16MB
)

init_db()


# ================================================================ SECURITY
# ---- CSRF protection (token per session, checked on every POST) ----
def _csrf_token():
    tok = session.get("_csrf")
    if not tok:
        tok = secrets.token_urlsafe(32)
        session["_csrf"] = tok
    return tok


@app.context_processor
def _inject_csrf():
    return {"csrf_token": _csrf_token}


@app.before_request
def _csrf_protect():
    if request.method == "POST":
        # allow the token-authenticated cron endpoint (no session/cookie)
        if request.path == "/cron/backup":
            return
        sent = request.form.get("_csrf") or request.headers.get("X-CSRF-Token")
        if not sent or not hmac.compare_digest(sent, session.get("_csrf", "")):
            # For JSON/API calls, return a clear 400 the frontend can handle.
            if request.path.startswith("/api/"):
                abort(400, "CSRF token missing or invalid.")
            # For normal form posts, don't dead-end the user on an error page.
            # If their session is gone, send them to log in; otherwise bounce
            # them back to the page they came from with a friendly message.
            if not session.get("user_id"):
                flash("Your session expired. Please log in again.", "err")
                return redirect(url_for("login"))
            flash("That form expired — please try again.", "err")
            ref = request.referrer
            if ref and urlparse(ref).netloc == urlparse(request.host_url).netloc:
                return redirect(ref)
            return redirect(url_for("home"))


# ---- login rate limiting (in-memory, per-IP) ----
_login_attempts = {}   # ip -> [timestamps]
LOGIN_MAX = 8          # attempts
LOGIN_WINDOW = 300     # seconds (5 min)


def _rate_limited(ip):
    now = time.time()
    hits = [t for t in _login_attempts.get(ip, []) if now - t < LOGIN_WINDOW]
    _login_attempts[ip] = hits
    return len(hits) >= LOGIN_MAX


def _record_attempt(ip):
    _login_attempts.setdefault(ip, []).append(time.time())


# ---- security headers on every response ----
@app.after_request
def _security_headers(resp):
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "DENY"                # no clickjacking
    resp.headers["Referrer-Policy"] = "same-origin"
    resp.headers["X-XSS-Protection"] = "1; mode=block"
    resp.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    # Content Security Policy: only allow resources from our own origin (no third party)
    resp.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "font-src 'self'; "
        "connect-src 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    if os.environ.get("HTTPS_ONLY", "1") == "1":
        resp.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    # cache static assets aggressively for speed
    if request.path.startswith("/static/"):
        resp.headers["Cache-Control"] = "public, max-age=604800"   # 7 days
    return resp


# ---------------------------------------------------------------- db per request
@app.before_request
def _open_db():
    g.db = get_conn()


@app.teardown_request
def _close_db(_=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


# make helpers available in every template
# bump this string whenever static files change to force browsers to reload them
ASSET_VER = "20260713h"


@app.context_processor
def inject_helpers():
    welcome = session.pop("just_logged_in", False)
    return {"user": current_user(), "can_access": can_access, "ASSET_VER": ASSET_VER,
            "show_welcome": welcome}


# ---------------------------------------------------------------- date ranges
def resolve_range(period, start, end):
    today = date.today()
    if period == "day":
        return today.isoformat(), today.isoformat()
    if period == "week":
        s = today - timedelta(days=today.weekday())
        return s.isoformat(), today.isoformat()
    if period == "month":
        return today.replace(day=1).isoformat(), today.isoformat()
    if period == "7d":
        return (today - timedelta(days=6)).isoformat(), today.isoformat()
    if period == "30d":
        return (today - timedelta(days=29)).isoformat(), today.isoformat()
    if period == "year":
        return today.replace(month=1, day=1).isoformat(), today.isoformat()
    if period == "custom" and start and end:
        return start, end
    return (today - timedelta(days=29)).isoformat(), today.isoformat()


def num(f, key, cast=float, default=0):
    v = (f.get(key) or "").strip()
    if v == "":
        return default
    try:
        return cast(v)
    except ValueError:
        return default


# ================================================================ AUTH
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "?").split(",")[0].strip()
        if _rate_limited(ip):
            flash("Too many login attempts. Please wait 5 minutes and try again.", "err")
            return render_template("login.html")
        u = request.form.get("username", "").strip()
        p = request.form.get("password", "")
        row = g.db.execute(
            "SELECT * FROM users WHERE username=? AND active=1", (u,)
        ).fetchone()
        if row and check_password_hash(row["password_hash"], p):
            session.clear()
            session["user_id"] = row["id"]
            session["username"] = row["username"]
            session["full_name"] = row["full_name"]
            session["role"] = row["role"]
            session["perms"] = load_perms(row)
            session.permanent = True
            session["just_logged_in"] = True
            return redirect(url_for("home"))
        _record_attempt(ip)
        flash("Invalid username or password.", "err")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ================================================================ DASHBOARDS
@app.route("/")
@login_required
def home():
    """Land on the best dashboard the user can see."""
    if can_access("dash_main"):
        return redirect(url_for("dashboard_main"))
    if can_access("dash_production"):
        return redirect(url_for("dashboard_production"))
    # no dashboard permission -> send to first allowed module or a notice
    for key in ("production", "procurement", "inventory", "sales", "finance", "masters"):
        if can_access(key):
            return redirect(url_for({
                "production": "production_list", "procurement": "procurement_list",
                "inventory": "inventory", "sales": "sales_list",
                "finance": "finance_page", "masters": "masters",
            }[key]))
    return render_template("no_access.html")


@app.route("/dashboard/main")
@require("dash_main")
def dashboard_main():
    return render_template("dashboard.html", variant="main")


@app.route("/dashboard/production")
@require("dash_production")
def dashboard_production():
    return render_template("dashboard.html", variant="production")


# ---- available metrics for custom widgets ----
# key -> (label, kind: 'timeseries' or 'products' or 'kpiset', needs_finance)
DASH_METRICS = [
    ("consumption",  "Electrical Consumption (Day vs Night)", "consumption", False),
    ("crushmerged",  "Crushing Production & Products (KG)",    "crushmerged",  False),
    ("cleanmerged",  "Cleaning Production & Products (KG)",    "cleanmerged",  False),
    ("inputoutput",  "Input vs Crushing Output (KG)",         "inputoutput",  False),
    ("crushcost",    "Crushing Cost / kg (₹)",                "crushcost",    True),
    ("waste",        "Waste (KG)",                            "series",       False),
    ("finance",      "Income vs Expense (₹)",                 "finance",      True),
]


def _default_layout(variant):
    base = ["consumption", "crushmerged", "cleanmerged", "inputoutput"]
    if variant == "main":
        base.insert(3, "finance")
        base.append("crushcost")
    widgets = []
    for m in base:
        widgets.append({"id": m, "metric": m, "type": "auto",
                        "size": "large" if m == "consumption" else "small",
                        "hidden": False})
    return widgets


@app.route("/api/dashboard/layout")
@login_required
def api_layout_get():
    variant = request.args.get("variant", "production")
    row = g.db.execute(
        "SELECT layout FROM dashboard_layouts WHERE user_id=? AND variant=?",
        (session["user_id"], variant)).fetchone()
    if row:
        try:
            layout = json.loads(row["layout"])
        except Exception:
            layout = _default_layout(variant)
    else:
        layout = _default_layout(variant)
    # metrics list (finance only on main dashboard AND if user can see money)
    allow_finance = (variant == "main") and can_access("dash_main")
    metrics = [dict(key=k, label=lb, kind=kd) for k, lb, kd, nf in DASH_METRICS
               if (not nf) or allow_finance]
    return jsonify(layout=layout, metrics=metrics)


@app.route("/api/dashboard/layout", methods=["POST"])
@login_required
def api_layout_save():
    variant = request.args.get("variant", "production")
    layout = request.get_json(force=True).get("layout", [])
    g.db.execute(
        """INSERT INTO dashboard_layouts (user_id, variant, layout, updated_at)
           VALUES (?,?,?,datetime('now'))
           ON CONFLICT(user_id, variant) DO UPDATE SET layout=excluded.layout, updated_at=datetime('now')""",
        (session["user_id"], variant, json.dumps(layout)))
    g.db.commit()
    return jsonify(ok=True)


# ---- live stock lookup for forms ----
@app.route("/api/stock")
@login_required
def api_stock():
    """Return current stock for an item. ?type=raw&name=Corn Cob  or ?type=raw&raw_id=1
    or ?type=finished&product_id=3 / &name=..."""
    db = g.db
    itype = request.args.get("type", "raw")
    name = request.args.get("name")
    if request.args.get("raw_id"):
        m = db.execute("SELECT name, low_stock FROM raw_materials WHERE id=?",
                       (request.args.get("raw_id"),)).fetchone()
        if m:
            name = m["name"]
    if request.args.get("product_id"):
        p = db.execute("SELECT name, low_stock FROM products WHERE id=?",
                       (request.args.get("product_id"),)).fetchone()
        if p:
            name = p["name"]
    if not name:
        return jsonify(qty=0, name=None)
    qty = logic.stock_for(db, itype, name)
    # threshold
    thr = 0
    if itype == "raw":
        row = db.execute("SELECT low_stock FROM raw_materials WHERE name=?", (name,)).fetchone()
    else:
        row = db.execute("SELECT low_stock FROM products WHERE name=?", (name,)).fetchone()
    if row:
        thr = row["low_stock"] or 0
    return jsonify(qty=qty, name=name, threshold=thr)


@app.route("/api/low_stock")
@login_required
def api_low_stock():
    return jsonify(alerts=logic.low_stock_alerts(g.db))


@app.route("/api/summary")
@login_required
def api_summary():
    db = g.db
    period = request.args.get("period", "30d")
    shift = request.args.get("shift", "combined")   # day | night | combined
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))

    if shift in ("Day", "day"):
        rows = db.execute(
            "SELECT * FROM reports WHERE report_date BETWEEN ? AND ? AND shift='Day' ORDER BY report_date, shift",
            (s, e)).fetchall()
    elif shift in ("Night", "night"):
        rows = db.execute(
            "SELECT * FROM reports WHERE report_date BETWEEN ? AND ? AND shift='Night' ORDER BY report_date, shift",
            (s, e)).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM reports WHERE report_date BETWEEN ? AND ? ORDER BY report_date, shift",
            (s, e)).fetchall()

    kpi = dict(consumption=0.0, crushing=0.0, cleaning=0.0, buckets=0,
               output=0.0, theli=0.0, waste=0.0, input_kg=0.0, reports=len(rows))
    by_date, shift_totals = {}, {"Day": 0.0, "Night": 0.0}
    for r in rows:
        kpi["consumption"] += r["consumption"] or 0
        kpi["crushing"] += r["crushing_total_kg"] or 0
        kpi["cleaning"] += r["cleaning_total_kg"] or 0
        kpi["buckets"] += r["raw_buckets"] or 0
        kpi["output"] += r["crushing_total_kg"] or 0
        kpi["theli"] += (r["crushing_total_theli"] or 0) + (r["cleaning_total_theli"] or 0)
        kpi["waste"] += r["waste_kg"] or 0
        row_input = (r["raw_buckets"] or 0) * (r["bucket_weight"] or 0)
        kpi["input_kg"] += row_input
        d = r["report_date"]
        rec = by_date.setdefault(d, dict(date=d, consumption=0, crushing=0,
                                         cleaning=0, buckets=0, output=0, waste=0, input_kg=0,
                                         day_consumption=0, night_consumption=0))
        rec["consumption"] += r["consumption"] or 0
        rec["crushing"] += r["crushing_total_kg"] or 0
        rec["cleaning"] += r["cleaning_total_kg"] or 0
        rec["buckets"] += r["raw_buckets"] or 0
        rec["output"] += r["crushing_total_kg"] or 0
        rec["waste"] += r["waste_kg"] or 0
        rec["input_kg"] += row_input
        if r["shift"] == "Day":
            rec["day_consumption"] += r["consumption"] or 0
        else:
            rec["night_consumption"] += r["consumption"] or 0
        shift_totals[r["shift"]] = shift_totals.get(r["shift"], 0) + (r["consumption"] or 0)

    total_prod = kpi["crushing"] + kpi["cleaning"]
    kpi["units_per_kg"] = round(kpi["consumption"] / total_prod, 4) if total_prod else 0
    # ---- minor efficiency calculations ----
    kpi["yield_pct"] = round(kpi["crushing"] / kpi["input_kg"] * 100, 1) if kpi["input_kg"] else 0
    kpi["waste_pct"] = round(kpi["waste"] / kpi["input_kg"] * 100, 1) if kpi["input_kg"] else 0
    kpi["output_per_unit"] = round(kpi["crushing"] / kpi["consumption"], 2) if kpi["consumption"] else 0
    # power cost per kg from the latest electricity bill covering this range's month
    try:
        month = e[:7]
        bill = db.execute("SELECT amount_inr FROM electricity_bills WHERE month=?", (month,)).fetchone()
        if bill and total_prod:
            kpi["power_cost_per_kg"] = round(bill["amount_inr"] / total_prod, 2)
        else:
            kpi["power_cost_per_kg"] = None
    except Exception:
        kpi["power_cost_per_kg"] = None
    for k in ("consumption", "crushing", "cleaning", "output", "theli", "waste", "input_kg"):
        kpi[k] = round(kpi[k], 2)

    series = [by_date[d] for d in sorted(by_date)]

    # product breakdown split by category
    prod = db.execute(
        """SELECT category, name, SUM(total_kg) kg FROM production_lines pl
           JOIN reports r ON r.id=pl.report_id
           WHERE r.report_date BETWEEN ? AND ?
           GROUP BY category, name ORDER BY kg DESC""",
        (s, e),
    ).fetchall()
    crushing_products = [dict(name=p["name"], kg=round(p["kg"] or 0, 2))
                         for p in prod if p["category"] == "crushing"]
    cleaning_products = [dict(name=p["name"], kg=round(p["kg"] or 0, 2))
                         for p in prod if p["category"] == "cleaning"]

    # per-day product breakdown (for merged stacked charts)
    daily = db.execute(
        """SELECT r.report_date d, pl.category cat, pl.name nm, SUM(pl.total_kg) kg
           FROM production_lines pl JOIN reports r ON r.id=pl.report_id
           WHERE r.report_date BETWEEN ? AND ?
           GROUP BY r.report_date, pl.category, pl.name""", (s, e)).fetchall()
    crush_names = [p["name"] for p in crushing_products]
    clean_names = [p["name"] for p in cleaning_products]
    dates_sorted = sorted(by_date)
    def build_stack(cat, names):
        # {product_name: [kg per date]}
        out = {n: [0] * len(dates_sorted) for n in names}
        di = {d: i for i, d in enumerate(dates_sorted)}
        for row in daily:
            if row["cat"] == cat and row["nm"] in out and row["d"] in di:
                out[row["nm"]][di[row["d"]]] = round(row["kg"] or 0, 2)
        return out
    crushing_stack = build_stack("crushing", crush_names)
    cleaning_stack = build_stack("cleaning", clean_names)

    # ---- previous-period comparison for trend badges ----
    from datetime import datetime as _dt
    try:
        sd = _dt.fromisoformat(s); ed = _dt.fromisoformat(e)
        span = (ed - sd).days + 1
        prev_e = (sd - timedelta(days=1)).isoformat()
        prev_s = (sd - timedelta(days=span)).isoformat()
        prows = db.execute(
            "SELECT * FROM reports WHERE report_date BETWEEN ? AND ?", (prev_s, prev_e)).fetchall()
        prev = dict(consumption=0.0, crushing=0.0, cleaning=0.0, input_kg=0.0, waste=0.0)
        for r in prows:
            prev["consumption"] += r["consumption"] or 0
            prev["crushing"] += r["crushing_total_kg"] or 0
            prev["cleaning"] += r["cleaning_total_kg"] or 0
            prev["input_kg"] += (r["raw_buckets"] or 0) * (r["bucket_weight"] or 0)
            prev["waste"] += r["waste_kg"] or 0

        def pct(cur, old):
            if not old:
                return None
            return round((cur - old) / old * 100, 1)
        trends = dict(
            consumption=pct(kpi["consumption"], prev["consumption"]),
            crushing=pct(kpi["crushing"], prev["crushing"]),
            cleaning=pct(kpi["cleaning"], prev["cleaning"]),
            input_kg=pct(kpi["input_kg"], prev["input_kg"]),
            waste=pct(kpi["waste"], prev["waste"]),
        )
    except Exception:
        trends = {}

    # ---- inventory value (finished stock × sell rate) ----
    rate_map2 = {p["name"]: (p["sell_rate"] or 0) for p in db.execute("SELECT name, sell_rate FROM products").fetchall()}
    _raw, _finished = logic.stock_levels(db)
    inv_value = round(sum((fi["qty"] or 0) * rate_map2.get(fi["item_name"], 0) for fi in _finished), 2)
    kpi["inventory_value"] = inv_value
    payload = dict(
        range=dict(start=s, end=e), kpi=kpi, series=series,
        shift_totals=shift_totals, trends=trends,
        crushing_products=crushing_products, cleaning_products=cleaning_products,
        crushing_stack=crushing_stack, cleaning_stack=cleaning_stack,
    )
    # ---- crushing cost-per-kg breakdown (ALL factory cost on crushing output) ----
    def cost_cat(cats):
        q = "SELECT COALESCE(SUM(amount_inr),0) v FROM finance WHERE direction='expense' AND entry_date BETWEEN ? AND ? AND category IN (%s)" % (
            ",".join("?" * len(cats)))
        return db.execute(q, (s, e, *cats)).fetchone()["v"]
    c_raw = cost_cat(["raw_material"])
    c_labour = cost_cat(["salary"])
    c_power = cost_cat(["electricity"])
    c_maint = cost_cat(["maintenance"])
    c_other = cost_cat(["freight", "other", "payment"])
    crush_kg = db.execute(
        "SELECT COALESCE(SUM(crushing_total_kg),0) v FROM reports WHERE report_date BETWEEN ? AND ?",
        (s, e)).fetchone()["v"]
    total_cost = c_raw + c_labour + c_power + c_maint + c_other
    cpk = round(total_cost / crush_kg, 2) if crush_kg else 0
    payload["crush_cost"] = dict(
        raw=round(c_raw, 2), labour=round(c_labour, 2), power=round(c_power, 2),
        maint=round(c_maint, 2), other=round(c_other, 2),
        total=round(total_cost, 2), output_kg=round(crush_kg, 1), cost_per_kg=cpk,
        # per-kg components
        pk_raw=round(c_raw / crush_kg, 2) if crush_kg else 0,
        pk_labour=round(c_labour / crush_kg, 2) if crush_kg else 0,
        pk_power=round(c_power / crush_kg, 2) if crush_kg else 0,
        pk_maint=round(c_maint / crush_kg, 2) if crush_kg else 0,
        pk_other=round(c_other / crush_kg, 2) if crush_kg else 0,
    )
    # monthly trend of crushing cost/kg (last 6 months)
    trend = []
    import calendar as _cal
    base = date.fromisoformat(e)
    for i in range(5, -1, -1):
        my = base.year + (base.month - 1 - i) // 12
        mm = (base.month - 1 - i) % 12 + 1
        ms = f"{my:04d}-{mm:02d}-01"
        me = f"{my:04d}-{mm:02d}-{_cal.monthrange(my, mm)[1]:02d}"
        tc = db.execute(
            "SELECT COALESCE(SUM(amount_inr),0) v FROM finance WHERE direction='expense' AND entry_date BETWEEN ? AND ?",
            (ms, me)).fetchone()["v"]
        tk = db.execute(
            "SELECT COALESCE(SUM(crushing_total_kg),0) v FROM reports WHERE report_date BETWEEN ? AND ?",
            (ms, me)).fetchone()["v"]
        trend.append(dict(month=f"{my:04d}-{mm:02d}", cost_per_kg=round(tc / tk, 2) if tk else 0))
    payload["crush_cost"]["trend"] = trend
    # only include money data if the user can see the Main dashboard
    if can_access("dash_main"):
        fin = logic.finance_summary(db, s, e)
        sales_inr = db.execute(
            "SELECT COALESCE(SUM(total_inr),0) FROM sales WHERE sale_date BETWEEN ? AND ?",
            (s, e)).fetchone()[0]
        export_inr = db.execute(
            "SELECT COALESCE(SUM(total_inr),0) FROM sales WHERE kind='export' AND sale_date BETWEEN ? AND ?",
            (s, e)).fetchone()[0]
        payload["finance"] = fin
        payload["sales_inr"] = round(sales_inr, 2)
        payload["export_inr"] = round(export_inr, 2)
    return jsonify(payload)


# ================================================================ PRODUCTION
@app.route("/production")
@login_required
def production_list():
    period = request.args.get("period", "30d")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    rows = g.db.execute(
        "SELECT * FROM reports WHERE report_date BETWEEN ? AND ? ORDER BY report_date DESC, shift",
        (s, e),
    ).fetchall()
    summ = dict(
        crushing=round(sum(r["crushing_total_kg"] or 0 for r in rows), 1),
        cleaning=round(sum(r["cleaning_total_kg"] or 0 for r in rows), 1),
        waste=round(sum(r["waste_kg"] or 0 for r in rows), 1),
        units=round(sum(r["consumption"] or 0 for r in rows), 1),
        days=len(rows),
    )
    return render_template("production_list.html", rows=rows, period=period, start=s, end=e, summ=summ)


@app.route("/production/entry", methods=["GET", "POST"])
@require("production")
def production_entry():
    db = g.db
    if request.method == "POST":
        f = request.form
        start_unit = num(f, "start_unit")
        close_unit = num(f, "close_unit")
        consumption = num(f, "consumption")
        if consumption == 0 and close_unit and close_unit >= start_unit:
            consumption = round(close_unit - start_unit, 2)

        cats = f.getlist("line_cat")
        names = f.getlist("line_name")
        thelis = f.getlist("line_theli")
        weights = f.getlist("line_weight")
        totals = f.getlist("line_total")

        def fnum(lst, i):
            try:
                return float(lst[i]) if lst[i].strip() else 0.0
            except (ValueError, IndexError):
                return 0.0

        ct = ck = lt = lk = 0.0
        rows = []
        for i in range(len(names)):
            nm = names[i].strip()
            if not nm:
                continue
            th, wt, tot = fnum(thelis, i), fnum(weights, i), fnum(totals, i)
            if tot == 0 and th and wt:
                tot = th * wt
            cat = cats[i] if i < len(cats) else "crushing"
            if cat == "crushing":
                ct += th; ck += tot
            else:
                lt += th; lk += tot
            rows.append((cat, nm, th, wt, tot))

        w_slno = f.getlist("worker_slno")
        w_name = f.getlist("worker_name")
        w_wid = f.getlist("worker_id")
        w_att = f.getlist("worker_attendance")
        w_hrs = f.getlist("worker_hours")
        w_ot = f.getlist("worker_ot")
        w_mach = f.getlist("worker_machine")
        wrows = []
        for i in range(len(w_name)):
            nm = w_name[i].strip()
            if nm:
                try:
                    sl = int(w_slno[i]) if w_slno[i].strip() else i + 1
                except (ValueError, IndexError):
                    sl = i + 1
                wid = None
                try:
                    wid = int(w_wid[i]) if i < len(w_wid) and w_wid[i].strip() else None
                except (ValueError, IndexError):
                    wid = None
                att = w_att[i] if i < len(w_att) else "present"
                try:
                    hrs = float(w_hrs[i]) if i < len(w_hrs) and w_hrs[i].strip() else 0
                except (ValueError, IndexError):
                    hrs = 0
                try:
                    ot = float(w_ot[i]) if i < len(w_ot) and w_ot[i].strip() else 0
                except (ValueError, IndexError):
                    ot = 0
                try:
                    mach = int(w_mach[i]) if i < len(w_mach) and w_mach[i].strip() else None
                except (ValueError, IndexError):
                    mach = None
                wrows.append((sl, nm, wid, att, hrs, ot, mach))

        waste_kg = num(f, "waste_kg")
        params = (
            f.get("report_date"), f.get("shift"),
            start_unit, close_unit, consumption,
            int(num(f, "persons_m1", int)), int(num(f, "persons_m2", int)),
            int(num(f, "reel", int)), f.get("on_time"), f.get("off_time"),
            int(num(f, "raw_buckets", int)), num(f, "bucket_weight", float, 25), num(f, "raw_output_kg"),
            round(ct, 2), round(ck, 2), round(lt, 2), round(lk, 2),
            waste_kg, f.get("light_gayi_time"),
            f.get("loading_powder"), f.get("loading_grit"), f.get("loading_bhunar"),
            f.get("maintenance"), num(f, "other_maint_cost", float, 0), int(num(f, "left_with_note", int)),
            f.get("half_attendance"), f.get("on_leave_names"),
            f.get("reporter"), f.get("office"), f.get("notes"),
            session.get("user_id"),
        )

        rid = f.get("report_id")
        if rid:
            db.execute(
                """UPDATE reports SET report_date=?,shift=?,start_unit=?,close_unit=?,
                consumption=?,persons_m1=?,persons_m2=?,reel=?,on_time=?,off_time=?,
                raw_buckets=?,bucket_weight=?,raw_output_kg=?,crushing_total_theli=?,crushing_total_kg=?,
                cleaning_total_theli=?,cleaning_total_kg=?,waste_kg=?,light_gayi_time=?,
                loading_powder=?,loading_grit=?,loading_bhunar=?,maintenance=?,other_maint_cost=?,
                left_with_note=?,half_attendance=?,on_leave_names=?,reporter=?,office=?,
                notes=?,created_by=? WHERE id=?""",
                params + (rid,),
            )
            db.execute("DELETE FROM production_lines WHERE report_id=?", (rid,))
            db.execute("DELETE FROM workers WHERE report_id=?", (rid,))
            logic.clear_source(db, "inventory_moves", "production", rid)
            logic.clear_source(db, "inventory_moves", "cleaning", rid)
            logic.clear_source(db, "finance", "machine", rid)
            logic.clear_source(db, "finance", "prodmaint", rid)
        else:
            cur = db.execute(
                """INSERT INTO reports (report_date,shift,start_unit,close_unit,consumption,
                persons_m1,persons_m2,reel,on_time,off_time,raw_buckets,bucket_weight,raw_output_kg,
                crushing_total_theli,crushing_total_kg,cleaning_total_theli,cleaning_total_kg,
                waste_kg,light_gayi_time,loading_powder,loading_grit,loading_bhunar,maintenance,other_maint_cost,
                left_with_note,half_attendance,on_leave_names,reporter,office,notes,created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                params,
            )
            rid = cur.lastrowid

        # product name -> product id map for inventory posting
        prod_map = {p["name"]: p["id"] for p in db.execute("SELECT id,name FROM products").fetchall()}
        for cat, nm, th, wt, tot in rows:
            db.execute(
                """INSERT INTO production_lines (report_id,category,name,theli,theli_weight,total_kg)
                   VALUES (?,?,?,?,?,?)""",
                (rid, cat, nm, th, wt, tot),
            )
            # AUTO inventory: add produced kg to finished stock (match on name if in catalog)
            if tot > 0:
                pid = prod_map.get(nm.split(" (")[0]) or prod_map.get(nm) or 0
                logic.post_inventory(db, f.get("report_date"), "finished", pid, nm,
                                     tot, "production", rid,
                                     "Auto from daily production",
                                     session.get("user_id"))
        for sl, nm, wid, att, hrs, ot, mach in wrows:
            db.execute("INSERT INTO workers (report_id,slno,name,worker_id,attendance,hours,ot_hours,machine_id) VALUES (?,?,?,?,?,?,?,?)",
                       (rid, sl, nm, wid, att, hrs, ot, mach))

        # AUTO inventory: cleaning consumes GRIT as its raw material (deduct from finished grit stock)
        grit_used = num(f, "cleaning_grit_used", float, 0)
        if grit_used > 0:
            # find the grit stock item name (the crushing "Ground Grit" product)
            grit_name = None
            for cat, nm, th, wt, tot in rows:
                low = nm.lower()
                if cat == "crushing" and ("ground grit" in low or "દળાયેલી" in nm):
                    grit_name = nm
                    break
            if not grit_name:
                # fall back to any existing grit stock item
                gm = db.execute(
                    """SELECT item_name FROM inventory_moves
                       WHERE item_type='finished' AND (LOWER(item_name) LIKE '%ground grit%' OR item_name LIKE '%દળાયેલી%')
                       ORDER BY id DESC LIMIT 1""").fetchone()
                grit_name = gm["item_name"] if gm else "દળાયેલી ગ્રીટ (Ground Grit)"
            grit_pid = prod_map.get(grit_name.split(" (")[0]) or prod_map.get(grit_name) or 0
            avail_grit = logic.stock_for(db, "finished", grit_name)
            if grit_used > avail_grit:
                flash(f"⚠ Stock warning: cleaning used {grit_used:.0f} kg of grit "
                      f"but only {avail_grit:.0f} kg was in stock. Grit inventory is now negative.", "warn")
            logic.post_inventory(db, f.get("report_date"), "finished", grit_pid, grit_name,
                                 -grit_used, "cleaning", rid,
                                 "Grit consumed by cleaning production",
                                 session.get("user_id"))

        # AUTO inventory: deduct raw material (input weight = buckets x bucket weight)
        input_kg = int(num(f, "raw_buckets", int)) * num(f, "bucket_weight", float, 25)
        raw_mat_id = num(f, "raw_material_id", int, None)
        if input_kg > 0:
            if not raw_mat_id:
                cc = db.execute("SELECT id FROM raw_materials WHERE name LIKE 'Corn Cob%' LIMIT 1").fetchone()
                raw_mat_id = cc["id"] if cc else 0
            rm = db.execute("SELECT name FROM raw_materials WHERE id=?", (raw_mat_id,)).fetchone()
            rm_name = rm["name"] if rm else "Raw material"
            # stock check BEFORE deducting (warn but allow)
            avail = logic.stock_for(db, "raw", rm_name)
            if input_kg > avail:
                flash(f"⚠ Stock warning: production used {input_kg:.0f} kg of {rm_name} "
                      f"but only {avail:.0f} kg was in stock. Inventory is now negative — "
                      f"please add a procurement entry.", "warn")
            logic.post_inventory(db, f.get("report_date"), "raw", raw_mat_id or 0, rm_name,
                                 -input_kg, "production", rid,
                                 f"Consumed in production ({int(num(f,'raw_buckets',int))} buckets)",
                                 session.get("user_id"))

        # waste ledger + inventory (out)
        if waste_kg > 0:
            db.execute(
                """INSERT INTO waste (waste_date,source,material,quantity_kg,created_by)
                   VALUES (?,?,?,?,?)""",
                (f.get("report_date"), "production", "shift waste", waste_kg,
                 session.get("user_id")),
            )
        # ---- #6: machine logs (per-machine output/units/labour/maintenance) ----
        db.execute("DELETE FROM machine_logs WHERE report_id=?", (rid,))
        m_id = f.getlist("machine_id")
        m_name = f.getlist("machine_name")
        m_out = f.getlist("machine_output")
        m_units = f.getlist("machine_units")
        m_maint = f.getlist("machine_maint")
        # labour per machine = wages of workers assigned to it (computed, not typed)
        # build a quick day-rate lookup for assigned workers
        wage_by_machine = {}
        for sl, nm, wid, att, hrs, ot, mach in wrows:
            if not mach or att == "absent":
                continue
            wm = db.execute("SELECT pay_type, pay_rate, ot_rate FROM worker_master WHERE id=?",
                            (wid,)).fetchone() if wid else None
            if not wm:
                continue
            rate = wm["pay_rate"] or 0
            otr = wm["ot_rate"] or 0
            if wm["pay_type"] == "monthly":
                day = rate / 30.0
            else:
                day = rate * (0.5 if att == "half" else 1)
            wage_by_machine[mach] = wage_by_machine.get(mach, 0) + day + (ot or 0) * otr
        for i in range(len(m_id)):
            try:
                mid = int(m_id[i]) if m_id[i].strip() else None
            except (ValueError, IndexError):
                mid = None
            if not mid:
                continue
            out = float(m_out[i]) if i < len(m_out) and m_out[i].strip() else 0
            un = float(m_units[i]) if i < len(m_units) and m_units[i].strip() else 0
            mnt = float(m_maint[i]) if i < len(m_maint) and m_maint[i].strip() else 0
            lab = round(wage_by_machine.get(mid, 0), 2)   # auto from assigned workers
            if out or un or lab or mnt:
                db.execute(
                    """INSERT INTO machine_logs (report_id,log_date,machine_id,machine_name,
                       output_kg,units,labour_cost,maint_cost) VALUES (?,?,?,?,?,?,?,?)""",
                    (rid, f.get("report_date"), mid,
                     m_name[i] if i < len(m_name) else "", out, un, lab, mnt))
                # only maintenance posts to finance here (labour already counted via payroll/attendance)
                if mnt > 0:
                    logic.post_finance(db, f.get("report_date"), "expense", "maintenance", mnt,
                                       description=f"Machine maintenance: {m_name[i] if i<len(m_name) else ''}",
                                       source="machine", ref_id=rid, user_id=session.get("user_id"))
        # other/general maintenance (free entry in Loading/Attendance section)
        other_maint = num(f, "other_maint_cost", float, 0)
        if other_maint > 0:
            logic.post_finance(db, f.get("report_date"), "expense", "maintenance", other_maint,
                               description=f"Other maintenance: {f.get('maintenance') or ''}"[:120],
                               source="prodmaint", ref_id=rid, user_id=session.get("user_id"))
        db.commit()
        flash("Production report saved (stock updated).", "ok")
        return redirect(url_for("production_list"))

    edit_id = request.args.get("edit")
    report = None
    lines = workers = mlogs = []
    if edit_id:
        report = db.execute("SELECT * FROM reports WHERE id=?", (edit_id,)).fetchone()
        lines = db.execute("SELECT * FROM production_lines WHERE report_id=?", (edit_id,)).fetchall()
        workers = db.execute("SELECT * FROM workers WHERE report_id=? ORDER BY slno", (edit_id,)).fetchall()
        mlogs = db.execute("SELECT * FROM machine_logs WHERE report_id=?", (edit_id,)).fetchall()
    lines_json = json.dumps([dict(category=l["category"], name=l["name"], theli=l["theli"],
                                  theli_weight=l["theli_weight"], total_kg=l["total_kg"]) for l in lines])
    workers_json = json.dumps([dict(slno=w["slno"], name=w["name"],
                                    worker_id=(w["worker_id"] if "worker_id" in w.keys() else None),
                                    attendance=(w["attendance"] if "attendance" in w.keys() else "present"),
                                    hours=(w["hours"] if "hours" in w.keys() else 0),
                                    ot_hours=(w["ot_hours"] if "ot_hours" in w.keys() else 0),
                                    machine_id=(w["machine_id"] if "machine_id" in w.keys() else None)) for w in workers])
    materials = db.execute("SELECT * FROM raw_materials WHERE active=1 ORDER BY name").fetchall()
    master_workers = db.execute("SELECT * FROM worker_master WHERE active=1 OR active IS NULL ORDER BY name").fetchall()
    master_workers_json = json.dumps([dict(id=w["id"], name=w["name"],
                                           default_hours=w["default_hours"]) for w in master_workers])
    master_machines = db.execute("SELECT * FROM machine_master WHERE active=1 OR active IS NULL ORDER BY name").fetchall()
    machines_json = json.dumps([dict(id=m["id"], name=m["name"], capacity_kg=m["capacity_kg"]) for m in master_machines])
    mlogs_json = json.dumps([dict(machine_id=x["machine_id"], machine_name=x["machine_name"],
                                  output_kg=x["output_kg"], units=x["units"],
                                  labour_cost=x["labour_cost"], maint_cost=x["maint_cost"]) for x in mlogs])
    # current grit stock (for cleaning raw-material reference)
    try:
        grit_stock = round(logic.stock_for(db, "finished", "દળાયેલી ગ્રીટ (Ground Grit)"), 1)
        if not grit_stock:
            gm = db.execute(
                """SELECT item_name FROM inventory_moves WHERE item_type='finished'
                   AND (LOWER(item_name) LIKE '%ground grit%' OR item_name LIKE '%દળાયેલી%')
                   ORDER BY id DESC LIMIT 1""").fetchone()
            if gm:
                grit_stock = round(logic.stock_for(db, "finished", gm["item_name"]), 1)
    except Exception:
        grit_stock = 0
    return render_template("production_entry.html", report=report,
                           materials=materials, master_workers=master_workers,
                           master_workers_json=master_workers_json,
                           machines_json=machines_json, mlogs_json=mlogs_json,
                           lines_json=lines_json, workers_json=workers_json,
                           grit_stock=grit_stock,
                           today=date.today().isoformat())


@app.route("/production/delete/<int:rid>", methods=["POST"])
@require("production")
def production_delete(rid):
    logic.clear_source(g.db, "inventory_moves", "production", rid)
    g.db.execute("DELETE FROM reports WHERE id=?", (rid,))
    g.db.commit()
    flash("Report deleted.", "ok")
    return redirect(url_for("production_list"))


# ================================================================ PROCUREMENT
@app.route("/procurement")
@require("procurement", write=False)
def procurement_list():
    period = request.args.get("period", "30d")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    rows = g.db.execute(
        """SELECT p.*, COALESCE(s.company_name,s.name) AS supplier_name, rm.name AS material_name
           FROM procurement p
           LEFT JOIN suppliers s ON s.id=p.supplier_id
           LEFT JOIN raw_materials rm ON rm.id=p.raw_material_id
           WHERE p.entry_date BETWEEN ? AND ?
           ORDER BY p.entry_date DESC, p.id DESC""",
        (s, e),
    ).fetchall()
    total = sum(r["total_cost"] or 0 for r in rows)
    total_qty = sum(r["quantity_kg"] or 0 for r in rows)
    total_paid = sum(r["paid"] or 0 for r in rows)
    summ = dict(total=round(total, 2), qty=round(total_qty, 1),
                paid=round(total_paid, 2), due=round(total - total_paid, 2), count=len(rows))
    return render_template("procurement.html", rows=rows, period=period,
                           start=s, end=e, total=round(total, 2), summ=summ)


@app.route("/procurement/entry", methods=["GET", "POST"])
@require("procurement")
def procurement_entry():
    db = g.db
    if request.method == "POST":
        f = request.form
        qty = num(f, "quantity_kg")
        rate = num(f, "rate_per_kg")
        freight = num(f, "freight_cost")
        total = round(qty * rate + freight, 2)
        rid = f.get("proc_id")

        # resolve/insert supplier by name if a new village supplier typed
        supplier_id = num(f, "supplier_id", int, None)
        material_id = num(f, "raw_material_id", int, None)

        params = (f.get("entry_date"), supplier_id, f.get("village"), material_id,
                  f.get("transport"), f.get("vehicle_no"), qty, rate, freight,
                  total, num(f, "paid"), f.get("notes"), session.get("user_id"))

        if rid:
            db.execute(
                """UPDATE procurement SET entry_date=?,supplier_id=?,village=?,raw_material_id=?,
                transport=?,vehicle_no=?,quantity_kg=?,rate_per_kg=?,freight_cost=?,total_cost=?,
                paid=?,notes=?,created_by=? WHERE id=?""", params + (rid,))
            logic.clear_source(db, "inventory_moves", "procurement", rid)
            logic.clear_source(db, "finance", "procurement", rid)
        else:
            cur = db.execute(
                """INSERT INTO procurement (entry_date,supplier_id,village,raw_material_id,
                transport,vehicle_no,quantity_kg,rate_per_kg,freight_cost,total_cost,paid,notes,created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""", params)
            rid = cur.lastrowid

        # auto inventory in + finance expense
        mat = db.execute("SELECT name FROM raw_materials WHERE id=?", (material_id,)).fetchone()
        mat_name = mat["name"] if mat else "Raw material"
        if qty > 0:
            logic.post_inventory(db, f.get("entry_date"), "raw", material_id or 0,
                                 mat_name, qty, "procurement", rid,
                                 f"From {f.get('village') or 'supplier'}", session.get("user_id"))
        goods_only = round(qty * rate, 2)
        if goods_only > 0:
            logic.post_finance(db, f.get("entry_date"), "expense", "raw_material",
                               goods_only, source="procurement", ref_id=rid,
                               description=f"{mat_name} {qty}kg", user_id=session.get("user_id"))
        if freight > 0:
            logic.post_finance(db, f.get("entry_date"), "expense", "freight",
                               freight, source="procurement", ref_id=rid,
                               description=f"Freight — {mat_name}", user_id=session.get("user_id"))
        db.commit()
        flash("Procurement saved (stock + expense posted).", "ok")
        return redirect(url_for("procurement_list"))

    edit_id = request.args.get("edit")
    row = db.execute("SELECT * FROM procurement WHERE id=?", (edit_id,)).fetchone() if edit_id else None
    suppliers = db.execute("SELECT * FROM suppliers ORDER BY name").fetchall()
    materials = db.execute("SELECT * FROM raw_materials WHERE active=1 ORDER BY name").fetchall()
    return render_template("procurement_entry.html", row=row, suppliers=suppliers,
                           materials=materials, today=date.today().isoformat(),
                           dial_codes=validators.DIAL_CODES)


@app.route("/procurement/delete/<int:rid>", methods=["POST"])
@require("procurement")
def procurement_delete(rid):
    logic.clear_source(g.db, "inventory_moves", "procurement", rid)
    logic.clear_source(g.db, "finance", "procurement", rid)
    g.db.execute("DELETE FROM procurement WHERE id=?", (rid,))
    g.db.commit()
    flash("Procurement deleted.", "ok")
    return redirect(url_for("procurement_list"))


# ================================================================ INVENTORY
@app.route("/inventory")
@require("inventory", write=False)
def inventory():
    db = g.db
    raw, finished = logic.stock_levels(db)
    # days of raw material left = current stock / avg daily consumption (last 30 days)
    from datetime import timedelta as _td
    s30 = (date.today() - _td(days=30)).isoformat()
    raw = [dict(r) for r in raw]
    for r in raw:
        used = db.execute(
            """SELECT COALESCE(SUM(ABS(qty_kg)),0) u FROM inventory_moves
               WHERE item_type='raw' AND item_name=? AND qty_kg<0 AND move_date>=?""",
            (r["item_name"], s30)).fetchone()["u"]
        avg_daily = used / 30 if used else 0
        r["days_left"] = round(r["qty"] / avg_daily, 1) if avg_daily > 0 and r.get("qty", 0) > 0 else None
    # attach sell_rate and compute value for finished goods
    rate_map = {p["name"]: (p["sell_rate"] or 0) for p in db.execute("SELECT name, sell_rate FROM products").fetchall()}
    finished = [dict(fi) for fi in finished]
    total_value = 0
    for fi in finished:
        rate = rate_map.get(fi["item_name"], 0)
        fi["rate"] = rate
        fi["value"] = round(fi["qty"] * rate, 2)
        total_value += fi["value"]
    total_value = round(total_value, 2)
    moves = db.execute(
        "SELECT * FROM inventory_moves ORDER BY move_date DESC, id DESC LIMIT 100"
    ).fetchall()
    return render_template("inventory.html", raw=raw, finished=finished, moves=moves,
                           total_value=total_value)


@app.route("/inventory/adjust", methods=["POST"])
@require("inventory")
def inventory_adjust():
    f = request.form
    qty = num(f, "qty_kg")
    if f.get("direction") == "out":
        qty = -abs(qty)
    logic.post_inventory(g.db, f.get("move_date") or date.today().isoformat(),
                         f.get("item_type"), num(f, "item_id", int, 0),
                         f.get("item_name"), qty, "manual", None,
                         f.get("note"), session.get("user_id"))
    g.db.commit()
    flash("Stock adjusted.", "ok")
    return redirect(url_for("inventory"))


# ================================================================ SALES / EXPORTS
@app.route("/sales")
@require("sales", write=False)
def sales_list():
    period = request.args.get("period", "30d")
    kind = request.args.get("kind", "all")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    q = """SELECT sa.*, COALESCE(c.company_name,c.name) AS customer_name, p.name AS product_name
           FROM sales sa LEFT JOIN customers c ON c.id=sa.customer_id
           LEFT JOIN products p ON p.id=sa.product_id
           WHERE sa.sale_date BETWEEN ? AND ?"""
    args = [s, e]
    if kind in ("domestic", "export"):
        q += " AND sa.kind=?"
        args.append(kind)
    q += " ORDER BY sa.sale_date DESC, sa.id DESC"
    rows = g.db.execute(q, args).fetchall()
    total_inr = sum(r["total_inr"] or 0 for r in rows)
    total_recv = sum(r["received"] or 0 for r in rows)
    total_qty = sum(r["quantity_kg"] or 0 for r in rows)
    summ = dict(total=round(total_inr, 2), received=round(total_recv, 2),
                pending=round(total_inr - total_recv, 2), qty=round(total_qty, 1), count=len(rows))
    return render_template("sales.html", rows=rows, period=period, start=s, end=e,
                           kind=kind, total_inr=round(total_inr, 2), summ=summ)


@app.route("/sales/entry", methods=["GET", "POST"])
@require("sales")
def sales_entry():
    db = g.db
    if request.method == "POST":
        f = request.form
        qty = num(f, "quantity_kg")
        rate = num(f, "rate")
        freight = num(f, "freight_cost")
        other = num(f, "other_cost")
        tax_pct = num(f, "tax_pct")
        currency = f.get("currency") or "INR"
        fx = num(f, "fx_rate", float, 1) or 1
        goods_cost = round(qty * rate, 2)                       # goods only
        tax_amount = round(goods_cost * tax_pct / 100, 2)       # tax on goods
        total_amt = round(goods_cost + tax_amount + freight + other, 2)  # final cost
        total_inr = round(total_amt * fx, 2)
        kind = f.get("kind") or "domestic"
        rid = f.get("sale_id")

        pid_val = num(f, "product_id", int, None)
        if not pid_val:   # 0 or None → inventory-only item, store NULL
            pid_val = None
        params = (f.get("sale_date"), kind, num(f, "customer_id", int, None),
                  pid_val, qty, rate,
                  goods_cost, tax_pct, tax_amount, currency, fx,
                  freight, other, total_amt, total_inr, num(f, "received"),
                  f.get("invoice_no"), f.get("vehicle_no"), f.get("port"), f.get("container_no"),
                  f.get("hs_code"), f.get("notes"), session.get("user_id"))

        if rid:
            db.execute(
                """UPDATE sales SET sale_date=?,kind=?,customer_id=?,product_id=?,quantity_kg=?,
                rate=?,goods_cost=?,tax_pct=?,tax_amount=?,currency=?,fx_rate=?,freight_cost=?,other_cost=?,
                total_amount=?,total_inr=?,received=?,invoice_no=?,vehicle_no=?,port=?,container_no=?,hs_code=?,
                notes=?,created_by=? WHERE id=?""",
                params + (rid,))
            logic.clear_source(db, "inventory_moves", "sale", rid)
            logic.clear_source(db, "finance", "sale", rid)
        else:
            cur = db.execute(
                """INSERT INTO sales (sale_date,kind,customer_id,product_id,quantity_kg,rate,
                goods_cost,tax_pct,tax_amount,currency,fx_rate,freight_cost,other_cost,total_amount,total_inr,received,
                invoice_no,vehicle_no,port,container_no,hs_code,notes,created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", params)
            rid = cur.lastrowid

        prod = db.execute("SELECT name FROM products WHERE id=?", (num(f, "product_id", int, 0),)).fetchone()
        # prefer the explicit product_name from the form (handles inventory-only items with id=0)
        pname = f.get("product_name") or (prod["name"] if prod else "Product")
        if qty > 0:
            avail = logic.stock_for(db, "finished", pname)
            if qty > avail:
                flash(f"⚠ Stock warning: sold {qty:.0f} kg of {pname} but only {avail:.0f} kg "
                      f"was in stock. Finished-goods inventory is now negative.", "warn")
            logic.post_inventory(db, f.get("sale_date"), "finished",
                                 num(f, "product_id", int, 0), pname, -qty,
                                 "sale", rid, f"Sold ({kind})", session.get("user_id"))
        if total_inr > 0:
            logic.post_finance(db, f.get("sale_date"), "income",
                               "export" if kind == "export" else "sales",
                               total_inr, currency, fx, total_amt, "sale", rid,
                               f"{pname} {qty}kg", session.get("user_id"))
        db.commit()
        flash("Sale saved (stock reduced, income posted).", "ok")
        return redirect(url_for("sales_list", kind=kind))

    edit_id = request.args.get("edit")
    row = db.execute("SELECT * FROM sales WHERE id=?", (edit_id,)).fetchone() if edit_id else None
    customers = db.execute("SELECT * FROM customers ORDER BY name").fetchall()
    # #11: show products WITH their current finished-stock, and include any
    # finished-inventory item even if it isn't in the products master yet
    prod_rows = db.execute("SELECT * FROM products WHERE active=1 ORDER BY name").fetchall()
    raw_stock, finished_stock = logic.stock_levels(db)
    stock_by_name = {s["item_name"]: s["qty"] for s in finished_stock}
    products = []
    seen = set()
    for p in prod_rows:
        d = dict(p)
        d["stock"] = round(stock_by_name.get(p["name"], 0), 1)
        products.append(d)
        seen.add(p["name"])
    # any finished-goods in inventory not in the master → add so they're sellable
    for name, qty in stock_by_name.items():
        if name not in seen and qty:
            products.append(dict(id=0, name=name, stock=round(qty, 1), unit="kg", _from_inventory=True))
    currencies = db.execute("SELECT * FROM currencies").fetchall()
    default_kind = request.args.get("kind", "domestic")
    return render_template("sales_entry.html", row=row, customers=customers,
                           products=products, currencies=currencies,
                           default_kind=default_kind, today=date.today().isoformat(),
                           dial_codes=validators.DIAL_CODES)


@app.route("/sales/delete/<int:rid>", methods=["POST"])
@require("sales")
def sales_delete(rid):
    logic.clear_source(g.db, "inventory_moves", "sale", rid)
    logic.clear_source(g.db, "finance", "sale", rid)
    g.db.execute("DELETE FROM sales WHERE id=?", (rid,))
    g.db.commit()
    flash("Sale deleted.", "ok")
    return redirect(url_for("sales_list"))


# ================================================================ WASTE
@app.route("/waste", methods=["GET", "POST"])
@require("waste", write=False)
def waste_page():
    db = g.db
    if request.method == "POST":
        if not can_access("waste", write=True):
            abort(403)
        f = request.form
        db.execute(
            """INSERT INTO waste (waste_date,source,material,quantity_kg,disposal,value_inr,notes,created_by)
               VALUES (?,?,?,?,?,?,?,?)""",
            (f.get("waste_date"), f.get("source"), f.get("material"),
             num(f, "quantity_kg"), f.get("disposal"), num(f, "value_inr"),
             f.get("notes"), session.get("user_id")))
        # recovered value as income
        val = num(f, "value_inr")
        if val > 0:
            logic.post_finance(db, f.get("waste_date"), "income", "waste_sale",
                               val, description=f"Waste sold: {f.get('material')}",
                               user_id=session.get("user_id"))
        db.commit()
        flash("Waste entry saved.", "ok")
        return redirect(url_for("waste_page"))

    period = request.args.get("period", "30d")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    rows = db.execute(
        "SELECT * FROM waste WHERE waste_date BETWEEN ? AND ? ORDER BY waste_date DESC, id DESC",
        (s, e)).fetchall()
    total_kg = sum(r["quantity_kg"] or 0 for r in rows)
    total_val = sum(r["value_inr"] or 0 for r in rows)
    # #12: average raw-material purchase price per kg across all procurement
    avg = db.execute(
        """SELECT COALESCE(SUM(quantity_kg*rate_per_kg),0) v, COALESCE(SUM(quantity_kg),0) q
           FROM procurement""").fetchone()
    avg_raw_rate = round(avg["v"] / avg["q"], 2) if avg["q"] else 0
    # attach computed cost per waste row
    rows2 = []
    for r in rows:
        d = dict(r)
        d["cost"] = round((r["quantity_kg"] or 0) * avg_raw_rate, 2)
        rows2.append(d)
    return render_template("waste.html", rows=rows2, period=period, start=s, end=e,
                           total_kg=round(total_kg, 2), total_val=round(total_val, 2),
                           avg_raw_rate=avg_raw_rate,
                           total_cost=round(total_kg * avg_raw_rate, 2))


@app.route("/waste/delete/<int:rid>", methods=["POST"])
@require("waste")
def waste_delete(rid):
    g.db.execute("DELETE FROM waste WHERE id=?", (rid,))
    g.db.commit()
    flash("Waste entry deleted.", "ok")
    return redirect(url_for("waste_page"))


# ================================================================ FINANCE
@app.route("/finance", methods=["GET", "POST"])
@require("finance")
def finance_page():
    db = g.db
    if request.method == "POST":
        f = request.form
        amt = num(f, "amount_inr")
        logic.post_finance(db, f.get("entry_date"), f.get("direction"),
                           f.get("category"), amt, description=f.get("description"),
                           user_id=session.get("user_id"))
        db.commit()
        flash("Finance entry saved.", "ok")
        return redirect(url_for("finance_page"))

    period = request.args.get("period", "month")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    summary = logic.finance_summary(db, s, e)
    rows = db.execute(
        "SELECT * FROM finance WHERE entry_date BETWEEN ? AND ? ORDER BY entry_date DESC, id DESC LIMIT 200",
        (s, e)).fetchall()
    # ---- derived minor metrics ----
    inc = summary.get("income", 0) or 0
    exp = summary.get("expense", 0) or 0
    metrics = {}
    metrics["expense_ratio"] = round(exp / inc * 100, 1) if inc else None
    # kg sold + sales value in range (finished goods leaving via sale)
    sold = db.execute(
        """SELECT COALESCE(SUM(ABS(qty_kg)),0) kg FROM inventory_moves
           WHERE item_type='finished' AND qty_kg<0 AND source='sale' AND move_date BETWEEN ? AND ?""",
        (s, e)).fetchone()["kg"]
    sales_val = db.execute(
        """SELECT COALESCE(SUM(amount_inr),0) v FROM finance
           WHERE direction='income' AND category IN ('sales','export') AND entry_date BETWEEN ? AND ?""",
        (s, e)).fetchone()["v"]
    metrics["avg_price_per_kg"] = round(sales_val / sold, 2) if sold else None
    # power cost per kg for the range's month, as a cost reference
    month = e[:7]
    bill = db.execute("SELECT amount_inr FROM electricity_bills WHERE month=?", (month,)).fetchone()
    prod = db.execute(
        """SELECT COALESCE(SUM(crushing_total_kg),0)+COALESCE(SUM(cleaning_total_kg),0) kg
           FROM reports WHERE report_date BETWEEN ? AND ?""", (s, e)).fetchone()["kg"]
    power_cpk = round(bill["amount_inr"] / prod, 2) if bill and prod else None
    if metrics["avg_price_per_kg"] is not None and power_cpk is not None:
        metrics["margin_per_kg"] = round(metrics["avg_price_per_kg"] - power_cpk, 2)
    else:
        metrics["margin_per_kg"] = None
    return render_template("finance.html", summary=summary, rows=rows,
                           metrics=metrics, period=period, start=s, end=e)


@app.route("/finance/delete/<int:rid>", methods=["POST"])
@require("finance")
def finance_delete(rid):
    row = g.db.execute("SELECT source FROM finance WHERE id=?", (rid,)).fetchone()
    if row and row["source"] not in (None, "manual"):
        flash("This entry was auto-posted from another page (e.g. sale/procurement/payroll). "
              "Edit it from its source page, not here.", "err")
        return redirect(url_for("finance_page"))
    g.db.execute("DELETE FROM finance WHERE id=?", (rid,))
    g.db.commit()
    flash("Finance entry deleted.", "ok")
    return redirect(url_for("finance_page"))


@app.route("/api/finance")
@require("finance")
def api_finance():
    period = request.args.get("period", "month")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    # monthly income vs expense trend for the range
    rows = g.db.execute(
        """SELECT substr(entry_date,1,7) ym, direction, SUM(amount_inr) amt
           FROM finance WHERE entry_date BETWEEN ? AND ?
           GROUP BY ym, direction ORDER BY ym""", (s, e)).fetchall()
    months = sorted({r["ym"] for r in rows})
    inc = {m: 0 for m in months}
    exp = {m: 0 for m in months}
    for r in rows:
        (inc if r["direction"] == "income" else exp)[r["ym"]] = round(r["amt"], 2)
    summary = logic.finance_summary(g.db, s, e)
    # expense breakdown by category (for pie/bar graph)
    catrows = g.db.execute(
        """SELECT COALESCE(category,'other') cat, SUM(amount_inr) amt
           FROM finance WHERE direction='expense' AND entry_date BETWEEN ? AND ?
           GROUP BY cat ORDER BY amt DESC""", (s, e)).fetchall()
    categories = [dict(name=r["cat"], amount=round(r["amt"], 2)) for r in catrows]
    return jsonify(dict(months=months,
                        income=[inc[m] for m in months],
                        expense=[exp[m] for m in months],
                        categories=categories,
                        summary=summary))


# ================================================================ MONTHLY EXCEL REPORT
@app.route("/report/excel")
@require("finance")
def report_excel():
    import openpyxl, io, calendar
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    db = g.db
    month = request.args.get("month") or date.today().strftime("%Y-%m")
    s = month + "-01"
    y, m = int(month[:4]), int(month[5:7])
    last = calendar.monthrange(y, m)[1]
    e = f"{month}-{last:02d}"

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="7C6CF0")
    title_font = Font(bold=True, size=14)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Sainath Agro Industries — Monthly Report {month}"
    ws["A1"].font = title_font
    summ = logic.finance_summary(db, s, e)
    prod = db.execute(
        """SELECT COALESCE(SUM(crushing_total_kg),0) cr, COALESCE(SUM(cleaning_total_kg),0) cl,
                  COALESCE(SUM(waste_kg),0) wa, COALESCE(SUM(consumption),0) un
           FROM reports WHERE report_date BETWEEN ? AND ?""", (s, e)).fetchone()
    sales_v = db.execute("SELECT COALESCE(SUM(total_inr),0) v FROM sales WHERE sale_date BETWEEN ? AND ?", (s, e)).fetchone()["v"]
    rows = [
        ("", ""),
        ("PRODUCTION", ""),
        ("Crushing output (kg)", round(prod["cr"], 1)),
        ("Cleaning output (kg)", round(prod["cl"], 1)),
        ("Waste (kg)", round(prod["wa"], 1)),
        ("Electricity units", round(prod["un"], 1)),
        ("", ""),
        ("MONEY", ""),
        ("Total sales (INR)", round(sales_v, 2)),
        ("Total income (INR)", round(summ.get("income", 0), 2)),
        ("Total expense (INR)", round(summ.get("expense", 0), 2)),
        ("Profit (INR)", round(summ.get("profit", 0), 2)),
    ]
    r = 3
    for label, val in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        if label in ("PRODUCTION", "MONEY"):
            ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16

    ws2 = wb.create_sheet("Sales")
    ws2.append(["Date", "Customer", "Product", "Qty kg", "Rate", "Total INR", "Received INR", "Vehicle"])
    style_header(ws2, 1, 8)
    for row in db.execute(
        """SELECT sa.sale_date, COALESCE(c.company_name,c.name) cust, p.name prod,
                  sa.quantity_kg, sa.rate, sa.total_inr, sa.received, sa.vehicle_no
           FROM sales sa LEFT JOIN customers c ON c.id=sa.customer_id
           LEFT JOIN products p ON p.id=sa.product_id
           WHERE sa.sale_date BETWEEN ? AND ? ORDER BY sa.sale_date""", (s, e)).fetchall():
        ws2.append([row["sale_date"], row["cust"] or "", row["prod"] or "", row["quantity_kg"],
                    row["rate"], row["total_inr"], row["received"], row["vehicle_no"] or ""])
    for col in "ABCDEFGH":
        ws2.column_dimensions[col].width = 14

    ws3 = wb.create_sheet("Expenses")
    ws3.append(["Date", "Category", "Amount INR", "Source", "Description"])
    style_header(ws3, 1, 5)
    for row in db.execute(
        """SELECT entry_date, category, amount_inr, source, description FROM finance
           WHERE direction='expense' AND entry_date BETWEEN ? AND ? ORDER BY entry_date""", (s, e)).fetchall():
        ws3.append([row["entry_date"], row["category"], row["amount_inr"], row["source"], row["description"] or ""])
    for col in "ABCDE":
        ws3.column_dimensions[col].width = 16

    ws4 = wb.create_sheet("Wages")
    ws4.append(["Worker", "Present", "Half", "OT hrs", "Base INR", "OT INR", "Total INR"])
    style_header(ws4, 1, 7)
    for wid, w in _compute_wages(db, s, e, month).items():
        ws4.append([w["name"], w["present"], w["half"], w["ot_hours"], w["base"], w["ot_pay"], w["total"]])
    for col in "ABCDEFG":
        ws4.column_dimensions[col].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"Sainath_Report_{month}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ================================================================ PARTY LEDGER (receivables / payables)
def _party_balances(db, party_type):
    """Compute outstanding balance per party.
    customer: sales total - received - payments_in  => they owe us (receivable)
    supplier: procurement total - paid - payments_out => we owe them (payable)"""
    rows = []
    if party_type == "customer":
        parties = db.execute("SELECT id, COALESCE(company_name,name) nm FROM customers ORDER BY nm").fetchall()
        for p in parties:
            billed = db.execute("SELECT COALESCE(SUM(total_inr),0) v FROM sales WHERE customer_id=?", (p["id"],)).fetchone()["v"]
            recv_inline = db.execute("SELECT COALESCE(SUM(received),0) v FROM sales WHERE customer_id=?", (p["id"],)).fetchone()["v"]
            pay = db.execute("SELECT COALESCE(SUM(amount_inr),0) v FROM payments WHERE party_type='customer' AND party_id=? AND direction='in'", (p["id"],)).fetchone()["v"]
            bal = round(billed - recv_inline - pay, 2)
            if billed or recv_inline or pay:
                rows.append(dict(id=p["id"], name=p["nm"], billed=round(billed, 2),
                                 settled=round(recv_inline + pay, 2), balance=bal))
    else:
        parties = db.execute("SELECT id, COALESCE(company_name,name) nm FROM suppliers ORDER BY nm").fetchall()
        for p in parties:
            billed = db.execute("SELECT COALESCE(SUM(total_cost),0) v FROM procurement WHERE supplier_id=?", (p["id"],)).fetchone()["v"]
            paid_inline = db.execute("SELECT COALESCE(SUM(paid),0) v FROM procurement WHERE supplier_id=?", (p["id"],)).fetchone()["v"]
            pay = db.execute("SELECT COALESCE(SUM(amount_inr),0) v FROM payments WHERE party_type='supplier' AND party_id=? AND direction='out'", (p["id"],)).fetchone()["v"]
            bal = round(billed - paid_inline - pay, 2)
            if billed or paid_inline or pay:
                rows.append(dict(id=p["id"], name=p["nm"], billed=round(billed, 2),
                                 settled=round(paid_inline + pay, 2), balance=bal))
    return rows


@app.route("/ledger", methods=["GET", "POST"])
@require("finance")
def ledger():
    db = g.db
    if request.method == "POST":
        f = request.form
        pt = f.get("party_type")
        pid = f.get("party_id")
        amt = num(f, "amount_inr", float, 0)
        direction = "in" if pt == "customer" else "out"
        if pid and amt > 0:
            nm = db.execute(
                f"SELECT COALESCE(company_name,name) nm FROM {'customers' if pt=='customer' else 'suppliers'} WHERE id=?",
                (pid,)).fetchone()
            db.execute(
                """INSERT INTO payments (pay_date, party_type, party_id, party_name, direction,
                   amount_inr, method, note, created_by) VALUES (?,?,?,?,?,?,?,?,?)""",
                (f.get("pay_date") or date.today().isoformat(), pt, pid,
                 nm["nm"] if nm else "", direction, amt, f.get("method"), f.get("note"),
                 session.get("user_id")))
            # payment received from customer = income; payment to supplier = expense
            if direction == "in":
                logic.post_finance(db, f.get("pay_date") or date.today().isoformat(), "income",
                                   "payment", amt, description=f"Payment from {nm['nm'] if nm else ''}",
                                   source="payment", user_id=session.get("user_id"))
            else:
                logic.post_finance(db, f.get("pay_date") or date.today().isoformat(), "expense",
                                   "payment", amt, description=f"Payment to {nm['nm'] if nm else ''}",
                                   source="payment", user_id=session.get("user_id"))
            db.commit()
            flash("Payment recorded.", "ok")
        return redirect(url_for("ledger", tab=f.get("tab", "customer")))

    tab = request.args.get("tab", "customer")
    customers = _party_balances(db, "customer")
    suppliers = _party_balances(db, "supplier")
    recv_total = round(sum(r["balance"] for r in customers if r["balance"] > 0), 2)
    pay_total = round(sum(r["balance"] for r in suppliers if r["balance"] > 0), 2)
    cust_list = db.execute("SELECT id, COALESCE(company_name,name) nm FROM customers ORDER BY nm").fetchall()
    supp_list = db.execute("SELECT id, COALESCE(company_name,name) nm FROM suppliers ORDER BY nm").fetchall()
    recent_pay = db.execute(
        "SELECT * FROM payments WHERE party_type=? ORDER BY pay_date DESC, id DESC LIMIT 20", (tab,)).fetchall()
    recent_pay = [dict(p) for p in recent_pay]
    return render_template("ledger.html", tab=tab, customers=customers, suppliers=suppliers,
                           recv_total=recv_total, pay_total=pay_total,
                           cust_list=cust_list, supp_list=supp_list, recent_pay=recent_pay)


# ================================================================ PAYROLL
@app.route("/payroll", methods=["GET", "POST"])
@require("finance")
def payroll():
    db = g.db
    month = request.args.get("month") or request.form.get("month") or date.today().strftime("%Y-%m")
    s = month + "-01"
    # month end
    y, m = int(month[:4]), int(month[5:7])
    if m == 12:
        e = f"{y}-12-31"
    else:
        import calendar
        last = calendar.monthrange(y, m)[1]
        e = f"{month}-{last:02d}"
    # date to post finance rows on: month-end, but never in the future
    post_date = min(e, date.today().isoformat())

    # POST = save a per-worker adjustment (edit)
    if request.method == "POST" and request.form.get("action") == "advance":
        wid = request.form.get("worker_id")
        kind = request.form.get("kind")  # 'given' or 'repaid'
        amt = num(request.form, "amount", float, 0)
        adate = request.form.get("adv_date") or date.today().isoformat()
        note = request.form.get("note")
        if wid and amt > 0 and kind in ("given", "repaid"):
            db.execute(
                """INSERT INTO worker_advances (worker_id, entry_date, kind, amount, note, created_by)
                   VALUES (?,?,?,?,?,?)""",
                (wid, adate, kind, amt, note, session.get("user_id")))
            # advance GIVEN = money out now → post to finance. repayment does NOT double-count.
            if kind == "given":
                wname = db.execute("SELECT name FROM worker_master WHERE id=?", (wid,)).fetchone()
                logic.post_finance(db, adate, "expense", "salary", amt,
                                   description=f"Advance (udhar) — {wname['name'] if wname else ''}",
                                   source="advance", ref_id=int(wid), user_id=session.get("user_id"))
            db.commit()
            flash("Advance recorded." if kind == "given" else "Repayment recorded.", "ok")
        return redirect(url_for("payroll", month=month))

    if request.method == "POST" and request.form.get("action") == "adjust":
        wid = request.form.get("worker_id")
        amt = num(request.form, "amount", float, 0)
        note = request.form.get("note")
        db.execute(
            """INSERT INTO wage_adjustments (month, worker_id, amount, note) VALUES (?,?,?,?)
               ON CONFLICT(month, worker_id) DO UPDATE SET amount=excluded.amount, note=excluded.note""",
            (month, wid, amt, note))
        db.commit()
        flash("Adjustment saved. Review and re-post to Finance.", "ok")
        return redirect(url_for("payroll", month=month))

    # POST = approve & post wages to finance
    if request.method == "POST" and request.form.get("action") == "post":
        # remove any prior wage postings for this month to avoid duplicates
        logic.clear_source(db, "finance", "payroll", month_ref(month))
        total = 0.0
        for wid, amt in _compute_wages(db, s, e, month).items():
            if amt["total"] > 0:
                logic.post_finance(db, post_date, "expense", "salary", amt["total"],
                                   description=f"Wages {month} — {amt['name']}",
                                   source="payroll", ref_id=month_ref(month),
                                   user_id=session.get("user_id"))
                total += amt["total"]
        db.commit()
        flash(f"Posted ₹{total:,.0f} wages for {month} to Finance.", "ok")
        return redirect(url_for("payroll", month=month))

    wages = _compute_wages(db, s, e, month)
    grand = round(sum(w["total"] for w in wages.values()), 2)
    already = db.execute(
        "SELECT COUNT(*) FROM finance WHERE source='payroll' AND ref_id=?",
        (month_ref(month),)).fetchone()[0]
    # ---- advance (udhar) ledger ----
    # this-month given/repaid per worker
    adv_month = {}
    for r in db.execute(
        """SELECT worker_id, kind, SUM(amount) amt FROM worker_advances
           WHERE entry_date BETWEEN ? AND ? GROUP BY worker_id, kind""", (s, e)).fetchall():
        adv_month.setdefault(r["worker_id"], {"given": 0, "repaid": 0})[r["kind"]] = r["amt"] or 0
    # running outstanding balance per worker (all-time up to end of month)
    adv_bal = {}
    for r in db.execute(
        """SELECT worker_id,
                  SUM(CASE WHEN kind='given' THEN amount ELSE -amount END) bal
           FROM worker_advances WHERE entry_date<=? GROUP BY worker_id""", (e,)).fetchall():
        adv_bal[r["worker_id"]] = round(r["bal"] or 0, 2)
    # merge advance data into each wage row and compute net payable
    total_out = 0.0
    for wid, w in wages.items():
        m = adv_month.get(wid, {"given": 0, "repaid": 0})
        w["udhar_taken"] = round(m.get("given", 0), 2)
        w["udhar_repaid"] = round(m.get("repaid", 0), 2)
        w["udhar_balance"] = adv_bal.get(wid, 0)   # running outstanding
        # net payable this month = earned − outstanding balance (capped at 0)
        w["net_payable"] = round(max(w["total"] - w["udhar_balance"], 0), 2)
        total_out += w["udhar_balance"] if w["udhar_balance"] > 0 else 0
    total_earned = grand
    total_advance_out = round(total_out, 2)
    total_net = round(sum(w["net_payable"] for w in wages.values()), 2)
    # machine labour for the month (entered on Daily Report per machine)
    mlabour = db.execute(
        """SELECT machine_name, SUM(labour_cost) lab FROM machine_logs
           WHERE log_date BETWEEN ? AND ? GROUP BY machine_name HAVING lab>0""",
        (s, e)).fetchall()
    machine_labour = [dict(name=r["machine_name"], labour=round(r["lab"], 2)) for r in mlabour]
    machine_labour_total = round(sum(m["labour"] for m in machine_labour), 2)
    return render_template("payroll.html", wages=list(wages.values()), month=month,
                           grand=grand, already_posted=already > 0,
                           machine_labour=machine_labour,
                           machine_labour_total=machine_labour_total,
                           total_earned=total_earned, total_advance_out=total_advance_out,
                           total_net=total_net)


def month_ref(month):
    """Turn 'YYYY-MM' into a stable integer ref id (YYYYMM)."""
    return int(month.replace("-", ""))


def _compute_wages(db, s, e, month=None):
    """Compute wages per worker from attendance between s and e.
    If month given, apply saved per-worker adjustments (bonus/deduction)."""
    workers = db.execute("SELECT * FROM worker_master").fetchall()
    adj = {}
    if month:
        for a in db.execute("SELECT worker_id, amount, note FROM wage_adjustments WHERE month=?", (month,)).fetchall():
            adj[a["worker_id"]] = dict(amount=a["amount"] or 0, note=a["note"])
    result = {}
    for w in workers:
        rows = db.execute(
            """SELECT wk.attendance, wk.hours, wk.ot_hours, wk.worker_id
               FROM workers wk JOIN reports r ON wk.report_id=r.id
               WHERE r.report_date BETWEEN ? AND ? AND wk.worker_id=?""",
            (s, e, w["id"])).fetchall()
        present = sum(1 for x in rows if x["attendance"] == "present")
        half = sum(1 for x in rows if x["attendance"] == "half")
        absent = sum(1 for x in rows if x["attendance"] == "absent")
        pay_type = w["pay_type"] if "pay_type" in w.keys() else "daily"
        rate = (w["pay_rate"] if "pay_rate" in w.keys() else 0) or 0
        ot_rate = (w["ot_rate"] if "ot_rate" in w.keys() else 0) or 0
        ot_hours = sum((x["ot_hours"] or 0) for x in rows)
        if pay_type == "monthly":
            base = rate
        else:
            base = rate * (present + 0.5 * half)
        ot_pay = ot_hours * ot_rate
        a = adj.get(w["id"], dict(amount=0, note=None))
        adjustment = a["amount"] or 0
        total = round(base + ot_pay + adjustment, 2)
        result[w["id"]] = dict(
            id=w["id"], name=w["name"], pay_type=pay_type, rate=rate,
            present=present, half=half, absent=absent, ot_hours=round(ot_hours, 1),
            base=round(base, 2), ot_pay=round(ot_pay, 2),
            adjustment=round(adjustment, 2), adj_note=a["note"], total=total)
    return result


# ================================================================ ELECTRICITY
@app.route("/electricity", methods=["GET", "POST"])
@require("finance")
def electricity():
    db = g.db
    if request.method == "POST":
        f = request.form
        action = f.get("action")
        # ---- daily solar entry ----
        if action == "solar_daily":
            d = f.get("solar_date") or date.today().isoformat()
            su = num(f, "solar_units", float, 0)
            db.execute(
                """INSERT INTO solar_daily (log_date, units, note) VALUES (?,?,?)
                   ON CONFLICT(log_date) DO UPDATE SET units=excluded.units, note=excluded.note""",
                (d, su, f.get("note")))
            db.commit()
            flash("Solar generation recorded.", "ok")
            return redirect(url_for("electricity"))
        # ---- monthly bill (with optional solar total) ----
        month = f.get("month")
        rate = num(f, "rate_per_unit", float, 0)
        units = num(f, "units", float, 0)
        solar = num(f, "solar_units_month", float, 0)
        amount = num(f, "amount_inr")
        if amount == 0 and units and rate:
            amount = round(units * rate, 2)
        db.execute(
            """INSERT INTO electricity_bills (month, amount_inr, units, rate_per_unit, solar_units, note) VALUES (?,?,?,?,?,?)
               ON CONFLICT(month) DO UPDATE SET amount_inr=excluded.amount_inr,
               units=excluded.units, rate_per_unit=excluded.rate_per_unit,
               solar_units=excluded.solar_units, note=excluded.note""",
            (month, amount, units, rate, solar, f.get("note")))
        logic.clear_source(db, "finance", "electricity", month_ref(month))
        elec_date = min(month + "-28", date.today().isoformat())
        logic.post_finance(db, elec_date, "expense", "electricity", amount,
                           description=f"MGVCL bill {month}", source="electricity",
                           ref_id=month_ref(month), user_id=session.get("user_id"))
        db.commit()
        flash("Electricity bill saved & posted to finance.", "ok")
        return redirect(url_for("electricity"))

    bills = db.execute("SELECT * FROM electricity_bills ORDER BY month DESC").fetchall()
    rows = []
    for b in bills:
        s = b["month"] + "-01"
        e = b["month"] + "-31"
        prod = db.execute(
            """SELECT COALESCE(SUM(crushing_total_kg),0)+COALESCE(SUM(cleaning_total_kg),0) kg
               FROM reports WHERE report_date BETWEEN ? AND ?""", (s, e)).fetchone()["kg"]
        cpk = round(b["amount_inr"] / prod, 3) if prod else None
        rate = b["rate_per_unit"] if "rate_per_unit" in b.keys() else 0
        solar = b["solar_units"] if "solar_units" in b.keys() else 0
        # sum daily solar for this month as a cross-check
        daily_solar = db.execute(
            "SELECT COALESCE(SUM(units),0) v FROM solar_daily WHERE log_date BETWEEN ? AND ?",
            (s, e)).fetchone()["v"]
        rows.append(dict(id=b["id"], month=b["month"], amount=b["amount_inr"], units=b["units"],
                         rate_per_unit=rate, solar_units=solar, daily_solar=round(daily_solar, 1),
                         production=round(prod, 1), cost_per_kg=cpk, note=b["note"]))
    # recent daily solar entries
    solar_days = db.execute("SELECT * FROM solar_daily ORDER BY log_date DESC LIMIT 30").fetchall()
    solar_days = [dict(d) for d in solar_days]
    e_summ = dict(
        total_bill=round(sum(r["amount"] for r in rows), 0),
        total_units=round(sum(r["units"] or 0 for r in rows), 0),
        total_solar=round(sum((r["solar_units"] or 0) or (r["daily_solar"] or 0) for r in rows), 0),
        months=len(rows))
    return render_template("electricity.html", rows=rows, solar_days=solar_days, e_summ=e_summ)


@app.route("/electricity/delete/<int:rid>", methods=["POST"])
@require("finance")
def electricity_delete(rid):
    db = g.db
    b = db.execute("SELECT month FROM electricity_bills WHERE id=?", (rid,)).fetchone()
    if b:
        logic.clear_source(db, "finance", "electricity", month_ref(b["month"]))
        db.execute("DELETE FROM electricity_bills WHERE id=?", (rid,))
        db.commit()
        flash("Electricity bill deleted.", "ok")
    return redirect(url_for("electricity"))


@app.route("/solar/delete/<int:rid>", methods=["POST"])
@require("finance")
def solar_delete(rid):
    g.db.execute("DELETE FROM solar_daily WHERE id=?", (rid,))
    g.db.commit()
    flash("Solar entry deleted.", "ok")
    return redirect(url_for("electricity"))


@app.route("/ledger/payment/delete/<int:rid>", methods=["POST"])
@require("finance")
def payment_delete(rid):
    db = g.db
    p = db.execute("SELECT * FROM payments WHERE id=?", (rid,)).fetchone()
    if p:
        # remove its finance posting (matched by source+ref won't work; match by note/amount/date)
        db.execute("""DELETE FROM finance WHERE source='payment' AND amount_inr=? AND entry_date=?
                      AND description LIKE ?""",
                   (p["amount_inr"], p["pay_date"], f"%{p['party_name']}%"))
        db.execute("DELETE FROM payments WHERE id=?", (rid,))
        db.commit()
        flash("Payment deleted.", "ok")
    return redirect(url_for("ledger", tab=request.args.get("tab", "customer")))


@app.route("/payroll/advance/delete/<int:rid>", methods=["POST"])
@require("finance")
def advance_delete(rid):
    db = g.db
    a = db.execute("SELECT * FROM worker_advances WHERE id=?", (rid,)).fetchone()
    if a:
        # if it was a 'given' advance, remove its finance posting too
        if a["kind"] == "given":
            db.execute("""DELETE FROM finance WHERE source='advance' AND ref_id=? AND amount_inr=?
                          AND entry_date=?""", (a["worker_id"], a["amount"], a["entry_date"]))
        db.execute("DELETE FROM worker_advances WHERE id=?", (rid,))
        db.commit()
        flash("Advance entry deleted.", "ok")
    return redirect(url_for("payroll", month=request.args.get("month")))


@app.route("/api/electricity_rate")
@require("finance")
def api_electricity_rate():
    """Return an approximate current MGVCL LT industrial rate as a suggestion.
    This is a reference only — the user should verify against their actual bill."""
    # GERC LT-MD / industrial effective ~₹ per unit incl. FPPPA (approx, 2026-27)
    # kept as a server-side constant we can update; not scraped live for reliability
    return jsonify(dict(rate=8.15, note="Approx. MGVCL LT industrial incl. FPPPA (~₹3.15) — verify with your bill",
                        source="GERC tariff 2026-27 (approximate)"))


# ================================================================ MASTERS (full CRUD)
def _custom_field_defs(db, entity):
    return db.execute(
        "SELECT * FROM custom_fields WHERE entity=? ORDER BY sort, id", (entity,)
    ).fetchall()


def _collect_custom(db, entity, f):
    """Read custom field values from form -> JSON string; validate required."""
    defs = _custom_field_defs(db, entity)
    vals, missing = {}, []
    for d in defs:
        v = (f.get("cf_" + d["field_key"]) or "").strip()
        if d["required"] and not v:
            missing.append(d["label"])
        vals[d["field_key"]] = v
    return json.dumps(vals), missing


def _save_supplier(db, f, sid=None):
    company = (f.get("company_name") or "").strip()
    ok, msg = validators.validate_required({"Company Name": company})
    if not ok:
        return False, msg
    okp, mp = validators.validate_phone(f.get("phone"))
    if not okp:
        return False, mp
    okg, mg = validators.validate_gst(f.get("gst_no"))
    if not okg:
        return False, mg
    custom, missing = _collect_custom(db, "supplier", f)
    if missing:
        return False, "Required custom fields: " + ", ".join(missing)
    # 'name' column kept = company for backwards compatibility / display
    fields = (company, company, f.get("village"), f.get("address"),
              f.get("contact_name"), f.get("phone_cc") or "+91", f.get("phone"),
              (f.get("gst_no") or "").upper(), f.get("email"), f.get("notes"), custom)
    if sid:
        db.execute("""UPDATE suppliers SET name=?,company_name=?,village=?,address=?,contact_name=?,
                   phone_cc=?,phone=?,gst_no=?,email=?,notes=?,custom=? WHERE id=?""", fields + (sid,))
        return True, sid
    cur = db.execute("""INSERT INTO suppliers (name,company_name,village,address,contact_name,
                   phone_cc,phone,gst_no,email,notes,custom) VALUES (?,?,?,?,?,?,?,?,?,?,?)""", fields)
    return True, cur.lastrowid


def _save_customer(db, f, cid=None):
    company = (f.get("company_name") or "").strip()
    ok, msg = validators.validate_required({"Company Name": company})
    if not ok:
        return False, msg
    okp, mp = validators.validate_phone(f.get("phone"))
    if not okp:
        return False, mp
    okg, mg = validators.validate_gst(f.get("gst_no"))
    if not okg:
        return False, mg
    custom, missing = _collect_custom(db, "customer", f)
    if missing:
        return False, "Required custom fields: " + ", ".join(missing)
    fields = (company, company, f.get("kind") or "domestic",
              f.get("address"), f.get("contact_name"), f.get("phone_cc") or "+91",
              f.get("phone"), (f.get("gst_no") or "").upper(), f.get("country"),
              f.get("city"), f.get("email"), f.get("vehicle_no"), f.get("notes"), custom)
    if cid:
        db.execute("""UPDATE customers SET name=?,company_name=?,kind=?,address=?,contact_name=?,
                   phone_cc=?,phone=?,gst_no=?,country=?,city=?,email=?,vehicle_no=?,notes=?,custom=? WHERE id=?""",
                   fields + (cid,))
        return True, cid
    cur = db.execute("""INSERT INTO customers (name,company_name,kind,address,contact_name,
                   phone_cc,phone,gst_no,country,city,email,vehicle_no,notes,custom)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", fields)
    return True, cur.lastrowid


@app.route("/masters", methods=["GET", "POST"])
@require("masters")
def masters():
    db = g.db
    if request.method == "POST":
        t = request.form.get("type")
        f = request.form
        rid = f.get("id") or None
        try:
            if t == "supplier":
                ok, res = _save_supplier(db, f, rid)
                if not ok:
                    flash(res, "err"); return redirect(url_for("masters", tab="suppliers"))
            elif t == "customer":
                ok, res = _save_customer(db, f, rid)
                if not ok:
                    flash(res, "err"); return redirect(url_for("masters", tab="customers"))
            elif t == "product":
                if rid:
                    db.execute("UPDATE products SET name=?,category=?,unit=?,low_stock=?,sell_rate=? WHERE id=?",
                               (f.get("name"), f.get("category"), f.get("unit") or "kg",
                                num(f, "low_stock", float, 0), num(f, "sell_rate", float, 0), rid))
                else:
                    db.execute("INSERT INTO products (name,category,unit,low_stock,sell_rate) VALUES (?,?,?,?,?)",
                               (f.get("name"), f.get("category"), f.get("unit") or "kg",
                                num(f, "low_stock", float, 0), num(f, "sell_rate", float, 0)))
            elif t == "raw":
                if rid:
                    db.execute("UPDATE raw_materials SET name=?,unit=?,low_stock=? WHERE id=?",
                               (f.get("name"), f.get("unit") or "kg", num(f, "low_stock", float, 0), rid))
                else:
                    db.execute("INSERT INTO raw_materials (name,unit,low_stock) VALUES (?,?,?)",
                               (f.get("name"), f.get("unit") or "kg", num(f, "low_stock", float, 0)))
            elif t == "machine":
                if rid:
                    db.execute("UPDATE machine_master SET name=?,capacity_kg=?,power_kw=?,note=? WHERE id=?",
                               (f.get("name"), num(f, "capacity_kg", float, 0),
                                num(f, "power_kw", float, 0), f.get("note"), rid))
                else:
                    db.execute("INSERT INTO machine_master (name,capacity_kg,power_kw,note,active) VALUES (?,?,?,?,1)",
                               (f.get("name"), num(f, "capacity_kg", float, 0),
                                num(f, "power_kw", float, 0), f.get("note")))
            elif t == "worker":
                if rid:
                    db.execute("""UPDATE worker_master SET name=?,phone=?,role=?,default_hours=?,
                               pay_type=?,pay_rate=?,ot_rate=? WHERE id=?""",
                               (f.get("name"), f.get("phone"), f.get("role"),
                                num(f, "default_hours", float, 12), f.get("pay_type") or "daily",
                                num(f, "pay_rate", float, 0), num(f, "ot_rate", float, 0), rid))
                else:
                    db.execute("""INSERT INTO worker_master (name,phone,role,default_hours,pay_type,pay_rate,ot_rate,active)
                               VALUES (?,?,?,?,?,?,?,1)""",
                               (f.get("name"), f.get("phone"), f.get("role"),
                                num(f, "default_hours", float, 12), f.get("pay_type") or "daily",
                                num(f, "pay_rate", float, 0), num(f, "ot_rate", float, 0)))
            elif t == "currency":
                db.execute("INSERT OR REPLACE INTO currencies (code,symbol,rate_to_inr) VALUES (?,?,?)",
                           ((f.get("code") or "").upper(), f.get("symbol"), num(f, "rate_to_inr", float, 1)))
            elif t == "custom_field":
                key = re.sub(r"[^a-z0-9_]", "_", (f.get("label") or "").lower()).strip("_")
                if rid:
                    db.execute("""UPDATE custom_fields SET entity=?,label=?,ftype=?,options=?,required=?,sort=?
                               WHERE id=?""",
                               (f.get("entity"), f.get("label"), f.get("ftype"),
                                f.get("options"), int(num(f, "required", int)), int(num(f, "sort", int)), rid))
                else:
                    db.execute("""INSERT OR REPLACE INTO custom_fields (entity,field_key,label,ftype,options,required,sort)
                               VALUES (?,?,?,?,?,?,?)""",
                               (f.get("entity"), key, f.get("label"), f.get("ftype"),
                                f.get("options"), int(num(f, "required", int)), int(num(f, "sort", int))))
            db.commit()
            flash("Saved.", "ok")
        except Exception as ex:
            flash(f"Error: {ex}", "err")
        return redirect(url_for("masters", tab=f.get("tab", "suppliers")))

    return render_template(
        "masters.html",
        suppliers=[dict(r) for r in db.execute("SELECT * FROM suppliers ORDER BY name").fetchall()],
        customers=[dict(r) for r in db.execute("SELECT * FROM customers ORDER BY name").fetchall()],
        products=[dict(r) for r in db.execute("SELECT * FROM products ORDER BY category,name").fetchall()],
        materials=[dict(r) for r in db.execute("SELECT * FROM raw_materials ORDER BY name").fetchall()],
        workers=[dict(r) for r in db.execute("SELECT * FROM worker_master ORDER BY name").fetchall()],
        machines=[dict(r) for r in db.execute("SELECT * FROM machine_master ORDER BY name").fetchall()],
        currencies=db.execute("SELECT * FROM currencies").fetchall(),
        custom_fields=db.execute("SELECT * FROM custom_fields ORDER BY entity,sort").fetchall(),
        cf_supplier=_custom_field_defs(db, "supplier"),
        cf_customer=_custom_field_defs(db, "customer"),
        dial_codes=validators.DIAL_CODES,
        active_tab=request.args.get("tab", "suppliers"),
    )


@app.route("/masters/delete", methods=["POST"])
@require("masters")
def masters_delete():
    db = g.db
    t = request.form.get("type")
    rid = request.form.get("id")
    tab = request.form.get("tab", "suppliers")

    # ---- #9: block deletion if the item is referenced anywhere ----
    def blocked(msg):
        flash(msg, "err")
        return redirect(url_for("masters", tab=tab))

    if t == "raw":
        row = db.execute("SELECT name FROM raw_materials WHERE id=?", (rid,)).fetchone()
        nm = row["name"] if row else None
        used = 0
        used += db.execute("SELECT COUNT(*) c FROM procurement WHERE raw_material_id=?", (rid,)).fetchone()["c"]
        if nm:
            used += db.execute("SELECT COUNT(*) c FROM inventory_moves WHERE item_type='raw' AND item_name=?",
                               (nm,)).fetchone()["c"]
        # reports reference raw material via raw_material_id if the column exists
        try:
            used += db.execute("SELECT COUNT(*) c FROM reports WHERE raw_material_id=?", (rid,)).fetchone()["c"]
        except Exception:
            pass
        if used:
            return blocked(f"Cannot delete '{nm}' — it is used in {used} record(s) "
                           f"(procurement/production/inventory). It stays to keep your records intact.")
    elif t == "product":
        row = db.execute("SELECT name FROM products WHERE id=?", (rid,)).fetchone()
        nm = row["name"] if row else None
        used = 0
        if nm:
            used += db.execute("SELECT COUNT(*) c FROM sales WHERE product_id=?", (rid,)).fetchone()["c"]
            used += db.execute("SELECT COUNT(*) c FROM inventory_moves WHERE item_type='finished' AND item_name=?",
                               (nm,)).fetchone()["c"]
        if used:
            return blocked(f"Cannot delete '{nm}' — it is used in {used} sale/inventory record(s).")
    elif t == "supplier":
        used = db.execute("SELECT COUNT(*) c FROM procurement WHERE supplier_id=?", (rid,)).fetchone()["c"]
        if used:
            return blocked(f"Cannot delete this supplier — used in {used} procurement record(s).")
    elif t == "customer":
        used = db.execute("SELECT COUNT(*) c FROM sales WHERE customer_id=?", (rid,)).fetchone()["c"]
        if used:
            return blocked(f"Cannot delete this customer — used in {used} sale(s).")
    elif t == "worker":
        used = db.execute("SELECT COUNT(*) c FROM workers WHERE worker_id=?", (rid,)).fetchone()["c"]
        if used:
            return blocked(f"Cannot delete this worker — they appear in {used} attendance record(s). "
                           f"You can leave them; inactive workers won't affect much.")

    elif t == "machine":
        used = db.execute("SELECT COUNT(*) c FROM machine_logs WHERE machine_id=?", (rid,)).fetchone()["c"]
        if used:
            return blocked(f"Cannot delete this machine — it has {used} production log(s).")

    table = {"supplier": "suppliers", "customer": "customers", "product": "products",
             "raw": "raw_materials", "custom_field": "custom_fields",
             "worker": "worker_master", "machine": "machine_master"}.get(t)
    if table:
        db.execute(f"DELETE FROM {table} WHERE id=?", (rid,))
        db.commit()
        flash("Deleted.", "ok")
    elif t == "currency":
        db.execute("DELETE FROM currencies WHERE code=?", (rid,))
        db.commit()
        flash("Deleted.", "ok")
    return redirect(url_for("masters", tab=tab))


# ---- JSON API for inline "+ Add" from other pages (supplier/customer) ----
@app.route("/api/quick_add/<entity>", methods=["POST"])
@login_required
def quick_add(entity):
    db = g.db
    f = request.form
    if entity == "supplier" and can_access("procurement"):
        ok, res = _save_supplier(db, f)
        if not ok:
            return jsonify(ok=False, error=res), 400
        db.commit()
        return jsonify(ok=True, id=res, name=f.get("company_name") or f.get("name"))
    if entity == "customer" and can_access("sales"):
        ok, res = _save_customer(db, f)
        if not ok:
            return jsonify(ok=False, error=res), 400
        db.commit()
        return jsonify(ok=True, id=res, name=f.get("company_name") or f.get("name"))
    if entity == "raw" and can_access("procurement"):
        cur = db.execute("INSERT INTO raw_materials (name,unit) VALUES (?,?)",
                         (f.get("name"), f.get("unit") or "kg"))
        db.commit()
        return jsonify(ok=True, id=cur.lastrowid, name=f.get("name"))
    if entity == "product" and can_access("sales"):
        cur = db.execute("INSERT INTO products (name,category,unit) VALUES (?,?,?)",
                         (f.get("name"), f.get("category") or "other", f.get("unit") or "kg"))
        db.commit()
        return jsonify(ok=True, id=cur.lastrowid, name=f.get("name"))
    return jsonify(ok=False, error="Not permitted"), 403


# ================================================================ USERS (owner only)
@app.route("/users", methods=["GET", "POST"])
@login_required
def users():
    if session.get("role") != "owner":
        flash("Only the Owner can manage users.", "err")
        return redirect(url_for("home"))
    db = g.db
    if request.method == "POST":
        f = request.form
        action = f.get("action")
        if action == "add":
            role = f.get("role")
            perms = f.getlist("perms")
            if not perms:
                perms = template_for(role)   # seed from role template if none ticked
            try:
                db.execute(
                    "INSERT INTO users (username,full_name,password_hash,role,permissions) VALUES (?,?,?,?,?)",
                    (f.get("username"), f.get("full_name"),
                     generate_password_hash(f.get("password")), role, json.dumps(perms)))
                db.commit()
                flash("User added.", "ok")
            except Exception:
                flash("Username already exists.", "err")
        elif action == "perms":
            uid = f.get("uid")
            perms = f.getlist("perms")
            db.execute("UPDATE users SET permissions=? WHERE id=?", (json.dumps(perms), uid))
            db.commit()
            # if editing own account, refresh session perms live
            if str(session.get("user_id")) == str(uid):
                session["perms"] = perms
            flash("Permissions updated.", "ok")
        elif action == "role":
            uid = f.get("uid")
            role = f.get("role")
            # apply role template as the new permission set
            db.execute("UPDATE users SET role=?, permissions=? WHERE id=?",
                       (role, json.dumps(template_for(role)), uid))
            db.commit()
            flash("Role updated (permissions reset to role template).", "ok")
        elif action == "toggle":
            db.execute("UPDATE users SET active=1-active WHERE id=?", (f.get("uid"),))
            db.commit()
        elif action == "reset":
            db.execute("UPDATE users SET password_hash=? WHERE id=?",
                       (generate_password_hash(f.get("password")), f.get("uid")))
            db.commit()
            flash("Password reset.", "ok")
        return redirect(url_for("users"))

    rows = db.execute("SELECT * FROM users ORDER BY role, username").fetchall()
    users_list = []
    for r in rows:
        d = dict(r)
        d["perm_list"] = load_perms(r)
        users_list.append(d)
    return render_template("users.html", rows=users_list, all_modules=ALL_MODULES,
                           role_templates={k: v for k, v in
                                           [(rn, template_for(rn)) for rn in
                                            ["owner", "manager", "accountant", "operator"]]})


# ================================================================ EXPORTS CSV
@app.route("/export/<what>.csv")
@login_required
def export_csv(what):
    db = g.db
    period = request.args.get("period", "30d")
    s, e = resolve_range(period, request.args.get("start"), request.args.get("end"))
    buf = io.StringIO()
    w = csv.writer(buf)
    if what == "production":
        w.writerow(["Date", "Shift", "Units", "Buckets", "Crushing KG",
                    "Cleaning KG", "Waste KG", "Reporter"])
        for r in db.execute("SELECT * FROM reports WHERE report_date BETWEEN ? AND ? ORDER BY report_date", (s, e)):
            w.writerow([r["report_date"], r["shift"], r["consumption"], r["raw_buckets"],
                        r["crushing_total_kg"], r["cleaning_total_kg"],
                        r["waste_kg"], r["reporter"]])
    elif what == "sales":
        w.writerow(["Date", "Kind", "Product", "Qty KG", "Rate", "Currency", "Total INR", "Invoice"])
        for r in db.execute("""SELECT sa.*,p.name pn FROM sales sa LEFT JOIN products p ON p.id=sa.product_id
                               WHERE sale_date BETWEEN ? AND ? ORDER BY sale_date""", (s, e)):
            w.writerow([r["sale_date"], r["kind"], r["pn"], r["quantity_kg"], r["rate"],
                        r["currency"], r["total_inr"], r["invoice_no"]])
    elif what == "finance":
        w.writerow(["Date", "Direction", "Category", "Amount INR", "Description"])
        for r in db.execute("SELECT * FROM finance WHERE entry_date BETWEEN ? AND ? ORDER BY entry_date", (s, e)):
            w.writerow([r["entry_date"], r["direction"], r["category"], r["amount_inr"], r["description"]])
    else:
        abort(404)
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={what}_{s}_to_{e}.csv"})


# ================================================================ BACKUP
@app.route("/backup")
@login_required
def backup_page():
    if session.get("role") != "owner":
        flash("Only the Owner can access backups.", "err")
        return redirect(url_for("home"))
    # list any saved backups on disk
    import os
    folder = os.environ.get("BACKUP_DIR") or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "backups")
    saved = []
    if os.path.isdir(folder):
        saved = sorted([x for x in os.listdir(folder) if x.endswith(".zip")], reverse=True)
    email_configured = bool(os.environ.get("BREVO_API_KEY") and os.environ.get("BACKUP_EMAIL_TO"))
    return render_template("backup.html", saved=saved,
                           email_configured=email_configured,
                           email_to=os.environ.get("BACKUP_EMAIL_TO", ""))


@app.route("/backup/download")
@login_required
def backup_download():
    if session.get("role") != "owner":
        abort(403)
    data, fname = backup_mod.make_backup_zip()
    return Response(data, mimetype="application/zip",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/backup/email", methods=["POST"])
@login_required
def backup_email_now():
    if session.get("role") != "owner":
        abort(403)
    ok, msg = backup_mod.email_backup()
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("backup_page"))


@app.route("/cron/backup")
def cron_backup():
    """Token-protected endpoint for an external scheduler (e.g. cron-job.org)
    to trigger the daily backup email. Call:
        https://<yoursite>/cron/backup?token=YOURTOKEN
    The token must match the BACKUP_CRON_TOKEN environment variable."""
    import os
    expected = os.environ.get("BACKUP_CRON_TOKEN", "")
    given = request.args.get("token", "")
    if not expected or given != expected:
        abort(403)
    # save to disk (best-effort) and email
    try:
        backup_mod.save_backup_to_disk()
    except Exception:
        pass
    ok, msg = backup_mod.email_backup()
    return jsonify(ok=ok, message=msg)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
